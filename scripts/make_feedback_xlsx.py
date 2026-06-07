"""
Generates Product_Feedback_Round2.xlsx at the project root.

Each round of product feedback gets its own sheet in the workbook, so the
deliverable carries a clear audit trail of issues raised and how each was
addressed. Re-runnable any time the wording needs tweaking.

Sheets:
  - Round 1 Feedback : the original "too technical" complaint that kicked off
                       the whole tone-rewrite arc
  - Round 2 Feedback : 4 items raised on the round-2 comparison docs
  - Round 3 Feedback : 4 items raised on the round-2 outputs after we
                       shipped the fixes (uppercase severity, self-contradicting
                       paragraphs, per-layer paragraphs louder than overall,
                       strong standalone conclusions)

Usage
─────
    python -m scripts.make_feedback_xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER = ["#", "Product feedback", "Where it appears", "Root cause", "Proposed fix"]

_ROUND1_ROWS: list[list[str]] = [
    [
        "1",
        "\"Appendix A depth difference: Impact on Results (from LLM) is too technical.\" "
        "The LLM-generated impact paragraphs in Appendix A read like climate-science peer "
        "literature — full of jargon and inaccessible to a business stakeholder audience.",
        "Appendix A — every layer's Impact on Results paragraph (DEM, TWI, Impervious, NDVI, "
        "NDBI, LST, LULC, Roads, Waterline).",
        "The Claude prompts in `_ask_claude_raster` and `_ask_claude_vector` explicitly requested "
        "\"professional formal tone\" and instructed the model to \"cite a specific metric and "
        "state its direct implication\" + \"explain the physical mechanism driving hazard risk.\" "
        "That wording pulled outputs toward research-paper voice and surfaced jargon like TWI, "
        "NDBI, evapotranspiration, albedo, pluvial, fluvial, overbank flow, etc. The "
        "`_LAYER_IMPACT_FALLBACK` strings used when Claude calls failed were equally technical, "
        "so the problem persisted even on failure.",
        "Coordinated rewrite of `scripts/ara_risk_insights.py`: "
        "(1) Added an explicit AUDIENCE: block to BOTH prompts naming the reader as a \"business "
        "stakeholder making decisions about the site, NOT a GIS or climate-science expert\" and "
        "explicitly listing forbidden jargon (NDVI, TWI, NDBI, evapotranspiration, albedo, pluvial, "
        "fluvial). "
        "(2) Rewrote sentence instructions: \"Cite a specific metric and state its direct "
        "implication\" → \"Explain in everyday terms what this measurement tells us about the "
        "site\"; \"Explain the physical mechanism driving hazard risk\" → \"State the practical "
        "implication for hazard risk at site in plain language.\" "
        "(3) Replaced \"Rules: professional formal tone\" with a bulleted list including \"Plain "
        "English; no specialist jargon\" + \"If a technical term is unavoidable, briefly define it "
        "inline\" + \"Tone: clear, accessible, professional. Like explaining to an executive, not "
        "a peer scientist.\" "
        "(4) Rewrote all 9 entries in `_LAYER_IMPACT_FALLBACK` from technical to plain-English "
        "equivalents — replaced phrases like \"topographic depressions\" / \"evapotranspiration "
        "and shading\" / \"overbank flow and pluvial accumulation\" with everyday language so "
        "the failure path is just as accessible as the success path.",
    ],
]

_ROUND2_ROWS: list[list[str]] = [
    [
        "1",
        "Decimal-place percentages in the AFTER paragraphs (e.g., 0.1%, 25.9%, 99.9%).",
        "Appendix A — Impact on Results paragraph for every layer where a percentage is cited.",
        "The prompt sends Claude percentages with one decimal place (`.1f` format) and the "
        "susceptibility-distribution JSON sent in the same prompt contains decimal `pct` values. "
        "Claude echoes them verbatim.",
        "1) New `_fmt_pct()` helper that renders whole numbers (e.g. 26%) and `<1%` for sub-1% "
        "non-zero values. 2) Apply at every percentage format site in both prompts. "
        "3) Pre-round the `pct` values inside the susceptibility-distribution JSON before sending. "
        "4) Add an explicit prompt rule: \"Express all percentages as whole numbers (e.g., 26%, "
        "not 25.9%). For values below 1%, write '<1%'. Never write decimal percentages.\"",
    ],
    [
        "2",
        "Informal language (\"actually get\" example) slipping into the new outputs.",
        "Spot-flagged by reviewer; not consistently seen in current outputs, but worth preventing "
        "drift in future regenerations.",
        "When de-jargoning the prompt last round, the \"plain English\" instruction didn't "
        "explicitly forbid informal phrasing. Claude has latitude to drift toward casual wording.",
        "Add a new rule to both prompts: \"Avoid informal phrasing ('actually', 'pretty', "
        "'kind of', 'a bit') and contractions ('don't', 'can't', 'we'd'). Professional, "
        "accessible prose only.\"",
    ],
    [
        "3",
        "Sentence 4 currently reads as \"the site team should consider…\" — too soft. "
        "Product wants back the older prescriptive style (\"Priority intervention should focus "
        "on…\") and explicitly NO references to \"site team\".",
        "Appendix A — Sentence 4 of every CRITICAL or ELEVATED severity layer.",
        "When de-jargoning, the `sentence_4_rule` was rewritten as \"the most important thing the "
        "site team should consider doing in response\" to feel more accessible. The change shifted "
        "the voice from prescriptive (\"do X\") to advisory (\"consider doing X\"), watering down "
        "actionability.",
        "Revert `sentence_4_rule` to prescriptive form: \"State the priority intervention needed "
        "to manage this risk. Be specific — name the action and the part of the site it applies "
        "to (e.g., 'Improve drainage along the southern low-elevation strip'). Do NOT use phrases "
        "like 'the site team should consider', 'we recommend', or 'consider evaluating'. Write the "
        "intervention as a direct, actionable instruction.\" Applied to both raster and vector "
        "prompts.",
    ],
    [
        "4",
        "\"Moderate\" used as a vague qualifier (e.g., \"moderate flood risk could be a concern\") "
        "without concrete backing.",
        "Appendix A — paragraphs across all severities, but most visible where severity is "
        "MODERATE or LOW.",
        "Claude leans on \"moderate\" as a hedge when input data isn't dramatic. The prompt "
        "doesn't forbid vague qualifiers, so it has license to use them.",
        "Add a new rule to both prompts: \"Avoid vague qualifiers like 'moderate', 'some', "
        "'a few' unless you immediately back them with a specific number, percentage, or area "
        "within the site. Prefer concrete description over hedging.\"",
    ],
]

_ROUND3_ROWS: list[list[str]] = [
    [
        "1",
        "Severity words (\"MODERATE\", \"ELEVATED\", \"LOW\") rendered in UPPERCASE in the AFTER "
        "prose. Reads as shouty / unnatural in a business report.",
        "Appendix A — every paragraph where the per-layer or overall severity is mentioned.",
        "`_severity_signal()` returns uppercase strings (`\"CRITICAL\"` / `\"ELEVATED\"` / "
        "`\"MODERATE\"` / `\"LOW\"`) and the prompt interpolated `{severity}` verbatim. Claude "
        "echoed the all-caps form into prose.",
        "1) Inject `{severity.lower()}` (now `severity_lc`) into the prompt so Claude sees "
        "lowercase. 2) Add an explicit rule to both prompts: \"Render severity words ('low', "
        "'moderate', 'elevated', 'critical') in lowercase / sentence case in prose. Never write "
        "them in all caps.\"",
    ],
    [
        "2",
        "Self-contradicting paragraphs. Shell Norco DEM AFTER text said \"the site is low-lying and "
        "close to ground level\" (sounds floodable) then concluded \"flood risk is low\" — the "
        "physical description and the verdict pulled in opposite directions.",
        "Appendix A — most visible on layers whose raw metric (low elevation, dense built-up, "
        "high LST) reads concerning but where the OVERALL site risk for that hazard is benign.",
        "The prompt had no awareness of the SITE's overall hazard risk. Claude reasoned only from "
        "the layer's metric in isolation, so it described the metric dramatically and then had to "
        "tack on a benign verdict — producing prose that worked against its own conclusion.",
        "1) New `_overall_hazard_profile()` helper in `scripts/ara_risk_insights.py` derives the "
        "site's overall flood/heat severity from the building-level risk counts already produced "
        "by `ara_exposure`. 2) Threaded into `site_info` and injected into both prompts as a new "
        "OVERALL SITE RISK field + profile block. 3) Added an explicit rule: \"If the layer's "
        "metric looks concerning but the overall risk is low or moderate, explain WHY in plain "
        "language — other factors (drainage, vegetation, terrain, distance from water, building "
        "stock) keep the overall risk in check.\"",
    ],
    [
        "3",
        "Per-layer paragraphs read more alarming than the site's overall risk. Shell Norco Heat: "
        "LULC, NDVI, and NDBI all generated \"ELEVATED\" / \"MODERATE\" language with phrases "
        "like \"increasing heat exposure for workers\", but the overall site heat risk is moderate "
        "(per ResSolv). Stacked together, the layers felt much worse than reality.",
        "Appendix A — all heat layers when overall is moderate; all flood layers when overall is "
        "low; per-layer Sentence 4 prescriptive interventions firing even when overall is benign.",
        "(a) The prompt had no OVERALL anchor, so each layer told its own story without coordination. "
        "(b) Sentence 4 (priority intervention) was gated on per-layer severity, so a single ELEVATED "
        "layer would produce an aggressive intervention even when the site overall was fine.",
        "1) The new OVERALL SITE RISK block (see fix #2) gives Claude one authoritative reference "
        "point to calibrate each layer against. 2) Re-gated Sentence 4 to fire on OVERALL severity, "
        "not per-layer — prescriptive interventions only appear when the site's overall flood/heat "
        "is elevated or critical. 3) Added an explicit calibration rule: \"Do NOT write conclusions "
        "stronger or more alarming than the overall site risk justifies.\"",
    ],
    [
        "4",
        "LLM drew strong standalone conclusions from a single layer's data without connecting back "
        "to the overall site picture. Reads as if each Appendix A layer were the whole story.",
        "Appendix A — every layer paragraph; problem compounds when 3+ layers all say something "
        "different from the overall verdict.",
        "The prompt's Sentence 3 told Claude to \"state the practical implication for {hazard} risk "
        "at {site_name}\" — phrased to let Claude declare a layer-level risk verdict on its own. "
        "Combined with no overall anchor, Claude invented per-layer conclusions.",
        "1) Sentence 3 rewritten: \"State the practical implication for {hazard} risk at "
        "{site_name} — calibrated to the OVERALL site risk ({overall}), not just this layer's "
        "metric.\" 2) Added explicit rule: \"Do NOT draw a standalone hazard verdict from one "
        "layer's metric. The susceptibility class and overall site risk are authoritative — use "
        "the metric to explain WHY, not to override the verdict.\"",
    ],
]


def _populate_sheet(ws, title: str, rows: list[list[str]]) -> None:
    """Populate one worksheet with a header row + body rows, styled consistently."""
    ws.title = title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    ws.append(_HEADER)
    for col_idx, _ in enumerate(_HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_alignment = Alignment(vertical="top", wrap_text=True)
    for row_data in rows:
        ws.append(row_data)
    for r in range(2, 2 + len(rows)):
        for c in range(1, len(_HEADER) + 1):
            ws.cell(row=r, column=c).alignment = body_alignment

    widths = {"A": 5, "B": 38, "C": 28, "D": 50, "E": 70}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    for r in range(2, 2 + len(rows)):
        ws.row_dimensions[r].height = 175

    ws.freeze_panes = "A2"


def build_workbook() -> Workbook:
    wb = Workbook()
    # Round 1 sheet (the original "too technical" complaint — uses the
    # workbook's default first sheet)
    _populate_sheet(wb.active, "Round 1 Feedback", _ROUND1_ROWS)
    # Round 2 sheet (appended)
    round2_ws = wb.create_sheet("Round 2 Feedback")
    _populate_sheet(round2_ws, "Round 2 Feedback", _ROUND2_ROWS)
    # Round 3 sheet (appended)
    round3_ws = wb.create_sheet("Round 3 Feedback")
    _populate_sheet(round3_ws, "Round 3 Feedback", _ROUND3_ROWS)
    return wb


def main() -> None:
    project_root = Path(__file__).resolve().parent.parent
    out_path = project_root / "Product_Feedback_Round2.xlsx"
    wb = build_workbook()
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
